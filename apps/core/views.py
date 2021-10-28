import django
from django.shortcuts import redirect, render
from django.http import HttpResponse
from django.urls import reverse
from django.views.generic import TemplateView
from django.views import View
from django.views.generic.edit import FormView
from django.contrib import messages
from . import forms, models
import logging
from openpyxl import Workbook
from django.conf import settings
import tempfile
from io import BytesIO

logger = logging.getLogger(__name__)

# Create your views here.


class Index(TemplateView):
    template_name = 'core/index.html'


class Passport(FormView):
    template_name = 'core/passport-form.html'
    form_class = forms.PassportForm

    def __init__(self, *args, **kwargs):
        self.member = None
        super().__init__(**kwargs)

    def form_valid(self, form: forms.PassportForm):
        member = form.save()
        messages.success(
            self.request, 'Your information was successfully uploaded')
        return redirect(reverse('core:passport-info-display', kwargs={'pk': member.pk}))

    def form_invalid(self, form):
        messages.error(self.request, "An error occured. please try again")
        return super().form_invalid(form)


class PassportInfo(View):
    def get(self, request, pk):
        member = models.Member.objects.get(pk=pk)
        return render(request, 'core/passport-form.html', {'member': member})


class MembersList(TemplateView):
    template_name = 'core/members-list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        members = models.Member.objects.all()
        context['members'] = sorted(members, key=lambda m: m.get_fullname())
        return context


def download_members_list(request):
    members_list = sorted(models.Member.objects.all(),
                          key=lambda m: m.get_fullname())
    members = [(
        idx + 1,
        member.lastname.upper(),
        member.firstname.upper(),
        member.position.upper(),
        member.get_team_display().upper(),
        member.passport.url) for idx, member in enumerate(members_list)]
    logger.debug(members)
    if request.GET.get("format") == "xlsx":
        wb=Workbook()
        sheet=wb.active
        sheet.title='Bio-Data'
        headers=['S/N',
                   'Lastname',
                   'Firstname',
                   'Position',
                   'Team',
                   'Passport']  # Headers if names are required to be discrete.
        headers_name_merged=['S/N',
                   'Fullname',
                   'Position',
                   'Team',
                   'Passport']  # Headers if names are required to be merged.

        # Logic to loop through `headers_name_merged` and `members_name_merged`
        if request.GET.get("names") == "merged":
            members_name_merged=[(
        idx + 1,
        member.get_fullname().upper(),
        member.position.upper(),
        member.get_team_display().upper(),
        member.passport.url) for idx, member in enumerate(members_list)]
            # to populate the excel sheet.
            for index, row in enumerate(sheet.iter_rows(min_row=1, max_col=5, max_row=len(members_name_merged) + 1)):
                if index == 0:
                    for idx, cell in enumerate(row):
                        cell.value=headers_name_merged[idx]

            for row in range(2, len(members_name_merged) + 2):
                for idx, col in enumerate(headers_name_merged):
                    sheet.cell(row=row, column=idx+1,
                               value=members_name_merged[row-2][idx])

        else:  # alternate logic if names are required to be discrete.
            for index, row in enumerate(sheet.iter_rows(min_row=1, max_col=6, max_row=len(members) + 1)):
                if index == 0:
                    for idx, cell in enumerate(row):
                        cell.value=headers[idx]

            for row in range(2, len(members) + 2):
                for idx, col in enumerate(headers):
                    sheet.cell(row=row, column=idx+1,
                               value=members[row-2][idx])

        # saving the file
        temp_file=BytesIO()
        wb.save(filename=temp_file)
        temp_file.seek(0)

        response=HttpResponse(
            content_type="application/vnd.ms-excel"
        )
        name = models.Site.objects.first().name
        response["Content-Disposition"]=f"inline; filename={name}-members-bio-data.xlsx"
        response["Content-Transfer-Encoding"]="binary"
        response.write(temp_file.read())
        temp_file.close()
        return response


def handler404(request, exception):
    logger.error("Could not resolve path: ", str(exception.get('path')))
    return render(request, 'core/404.html')


def handler500(request):
    return render(request, 'core/500.html')
